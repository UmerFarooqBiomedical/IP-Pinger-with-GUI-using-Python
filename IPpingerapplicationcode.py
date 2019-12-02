# -*- coding: utf-8 -*-
"""
Created on Wed Nov 25 15:46:14 2019

@author: umer farooq

IP Pinger Application
"""
import tkinter as tk
from tkinter import ttk
import openpyxl

LARGE_FONT= ("Verdana bold", 12)
NORM_FONT= ("Verdana", 10)
SMALL_FONT= ("Verdana", 8)

class  IPping(tk.Tk):
    def __init__(self, *args, **kwargs):
        
        tk.Tk.__init__(self, *args, **kwargs)
        container = tk.Frame(self)
        tk.Tk.wm_title(self,  "IP Pinger")
        container.pack(side="top", fill="both", expand = True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        self.frames = {}

        for F in (HomePage, PageOne):

            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")
        self.show_frame(HomePage)

    def show_frame(self, cont):

        frame = self.frames[cont]
        frame.tkraise()
        


class HomePage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        tk.Frame.config(self, bg='#A0DEDC')
        label = tk.Label(self,  font=LARGE_FONT)
        label.config(text="IP Pinger", bg = '#A0DEDC')
        label.pack(pady=10,padx=10)
        label = tk.Label(self,  font=SMALL_FONT)
        label.config(text="""This tool can parse an excel sheet and perform 
a constant ping of IP addresses
in the spreadsheet.""", bg = '#CBF7F6')
        label.pack(pady=10,padx=10)
        
        button = ttk.Button(self, text="Start",
                            command=lambda: controller.show_frame(PageOne))
        button.pack()
        
        label2 = tk.Label(self)
        label2.config(text="""""", bg = '#A0DEDC')
        label2.pack()

#For first IP Address Change column
def writeIP1(self,FileName):
    print("Value of LOC4: %r: " % self.loc)
    wbkName = FileName      
    print('Submitted IP: {}'.format(self.text_box3.get("1.0", "end-1c")))
    NewValue1 = self.text_box3.get("1.0", "end-1c")
    print("New Value 1 = %r" % NewValue1)
    wbk = openpyxl.load_workbook(wbkName)
    wks = wbk['Sheet1']  #Name of the sheet in Excel file which you want to work on
    wks.cell((self.loc)+1, column=23+1).value = NewValue1 #You can change value 23 to any number of column you want to change
    wbk.save(wbkName)
    wbk.close
    print("IP1 changed")

#For second IP Address Change column
def writeIP2(self,FileName):
    print("Value of LOC5: %r: " % self.loc)
    wbkName = FileName  
    print('Submitted IP: {}'.format(self.text_box4.get("1.0", "end-1c")))
    NewValue2 = self.text_box4.get("1.0", "end-1c")
    wbk = openpyxl.load_workbook(wbkName)
    wks = wbk['Sheet1']  #Name of the sheet in Excel file which you want to work on
    wks.cell(self.loc+1, column=24+1).value = NewValue2 #You can change value 24 to any number of column you want to change
    wbk.save(wbkName)
    wbk.close
    print("IP2 changed")    

#For third IP Address Change column
def writeIP3(self, FileName):
    wbkName = FileName
    print('Submitted IP: {}'.format(self.text_box5.get("1.0", "end-1c")))
    NewValue3 = self.text_box5.get("1.0", "end-1c")
    wbk = openpyxl.load_workbook(wbkName)
    wks = wbk['Sheet1']  #Name of the sheet in Excel file which you want to work on
    wks.cell(self.loc+1, column=25+1).value = NewValue3 #You can change value 25 to any number of column you want to change
    wbk.save(wbkName)
    wbk.close
    print("IP3 changed")
    
def popupmsg(self, msg):
    popup = tk.Tk()
    popup.wm_title("Assigned IP Check Window")
    print('Submitted File Name: {}'.format(self.text_box1.get("1.0", "end-1c")))
    FileName = self.text_box1.get("1.0", "end-1c")
    print('Entered Location: {}'.format(self.text_box2.get("1.0", "end-1c")))
    Location = str(self.text_box2.get("1.0", "end-1c"))
    print(Location)

    from xlrd import open_workbook
    #Method to read from Excel file
    book = open_workbook(FileName)
    sheet = book.sheet_by_index(0)
    
    for r in range(sheet.nrows):
                for c in range(sheet.ncols):
                    cell = sheet.cell(r,c)
                    if cell.value == Location:
                       self.loc = r #index of interesting row
                       
    loc = self.loc
    cell1 = sheet.cell(loc, 23) #You can change value 23 to any number of column you want to change
    cell2 = sheet.cell(loc, 24) #You can change value 24 to any number of column you want to change
    cell3 = sheet.cell(loc, 25) #You can change value 25 to any number of column you want to change
    
    print ("IP Address for Reg is %r" % cell1.value)
    print ("IP Address for PC is %r" % cell2.value)
    print ("IP Address for Wifi is %r" % cell3.value)    
    label = ttk.Label(popup, text=msg, font=NORM_FONT)
    label.pack(side="top", fill="x", pady=10)
    
    label = ttk.Label(popup, text="IP Address of REG: %r" % cell1.value, font=NORM_FONT)
    label.pack()
    
    label = ttk.Label(popup, text="IP Address of PC: %r" % cell2.value, font=NORM_FONT)
    label.pack()
    
    label = ttk.Label(popup, text= "IP Address of Wifi:%r" % cell3.value, font=NORM_FONT)
    label.pack()
    print("Value of LOC1: %r: " % loc)
    
    B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
    B1.pack()
    print("Value of LOC2: %r: " % loc)
    popup.mainloop()
    print("Value of LOC3: %r: " % loc)
    
def print_something1(self):
           
    print('Submitted Time: {}'.format(self.text_box1.get("1.0", "end-1c")))
    FileName = self.text_box1.get("1.0", "end-1c")
    print(FileName)
class PageOne(tk.Frame):

    
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        label = tk.Label(self,  font=SMALL_FONT)
        label.config(text="""Enter the Excel Sheet name which you want to open.""", bg = '#F6ECEA')
        label.pack(pady=10,padx=10)
        tk.Frame.config(self, bg='#DFBFB8') 
        
        label3 = tk.Label(self)
        label3.config(text="""Enter Filename e.g File1.xlsx""", bg = '#DFBFB8')
        label3.pack()
            
        self.text_box1 = tk.Text(self, height=1, width=10)
        self.text_box1.pack(ipadx=80,ipady=2)
        
        button3 = ttk.Button(self, text="Submit",
                            command=lambda: (print_something1(self) ))
        button3.pack(pady = 4)
              
        label8 = tk.Label(self,  font=SMALL_FONT)
        label8.config(text="""Enter the machine location, which you want to Ping.""", bg = '#F6ECEA')
        label8.pack(pady=10,padx=10)
        tk.Frame.config(self, bg='#DFBFB8')
        
        label4 = tk.Label(self)
        label4.config(text="""Enter Machine Location e.g 1219b""", bg = '#DFBFB8')
        label4.pack()
        
        self.text_box2 = tk.Text(self, height=1, width=10)
        self.text_box2.pack(ipadx=40)
        
        
        button4 = ttk.Button(self, text="Ping",
                            command=lambda: (popupmsg(self, "Take a look into assigned IPs") ))
        button4.pack(pady = 4)
        
        label8 = tk.Label(self,  font=SMALL_FONT)
        label8.config(text="""Change the IP address using the slots below.""", bg = '#F6ECEA')
        label8.pack(pady=10,padx=10)
        tk.Frame.config(self, bg='#DFBFB8')
                        

        label11 = tk.Label(self)
        label11.config(text="""Change IP Address for Reg""", bg = '#DFBFB8')
        label11.pack()
            
        self.text_box3 = tk.Text(self, height=1, width=10)
        self.text_box3.pack(ipadx=40)
        
        button5 = ttk.Button(self, text="Change",
                            command=lambda: (self, writeIP1(self, FileName = self.text_box1.get("1.0", "end-1c")) ))
        button5.pack(pady = 4)
        
        label15 = tk.Label(self)
        label15.config(text="""Change IP Address for PC""", bg = '#DFBFB8')
        label15.pack()
            
        self.text_box4 = tk.Text(self, height=1, width=10)
        self.text_box4.pack(ipadx=40)
        
        button6 = ttk.Button(self, text="Change",
                            command=lambda: (writeIP2(self,FileName = self.text_box1.get("1.0", "end-1c")) ))
        button6.pack(pady = 4)
        
        label18 = tk.Label(self)
        label18.config(text="""Change IP Address for Wifi""", bg = '#DFBFB8')
        label18.pack()
            
        self.text_box5 = tk.Text(self, height=1, width=10)
        self.text_box5.pack(ipadx=40)
        
        button7 = ttk.Button(self, text="Change",
                            command=lambda: (writeIP3(self,FileName = self.text_box1.get("1.0", "end-1c")) )) 
        button7.pack(pady = 4)


        button1 = ttk.Button(self, text="Back to Home",
                            command=lambda:  controller.show_frame(HomePage))
        button1.pack(pady = 8)
        
app = IPping()
app.geometry("600x600") #Can increase or decrease default size of window by changing pixel values.
app.mainloop()
